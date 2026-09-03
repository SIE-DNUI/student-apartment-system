# -*- coding: utf-8 -*-
"""
房间管理路由模块
提供房间信息管理功能
"""
from flask import render_template, Blueprint, redirect, url_for, flash, request, send_file
from flask_login import login_required
from flask_wtf import FlaskForm
from wtforms import StringField, SelectField, IntegerField
from wtforms.validators import DataRequired, Optional, NumberRange
from app.models import db
from app.models import Room, FeeStandard, Student
from app.decorators import permission_required
from app.floor_layouts import LAYOUTS, detect_tower, get_layout
import io

bp = Blueprint('rooms', __name__, url_prefix='/rooms')


class RoomForm(FlaskForm):
    building = StringField('楼号', validators=[DataRequired(message='请输入楼号')])
    room_number = StringField('房间号', validators=[DataRequired(message='请输入房间号')])
    capacity = IntegerField('房间容量', validators=[DataRequired(), NumberRange(min=1, max=10)])
    floor = IntegerField('楼层', validators=[Optional()])
    fee_standard_id = SelectField('收费标准', coerce=int, validators=[Optional()])
    description = StringField('备注', validators=[Optional()])


@bp.route('/')
@login_required
def index():
    """房间列表"""
    page = request.args.get('page', 1, type=int)
    per_page = 30
    
    building_filter = request.args.get('building', '')
    status_filter = request.args.get('status', '')
    
    query = Room.query
    
    if building_filter:
        query = query.filter(Room.building == building_filter)
    
    if status_filter:
        if status_filter == 'available':
            query = query.filter(Room.current_occupancy < Room.capacity)
        elif status_filter == 'full':
            query = query.filter(Room.current_occupancy >= Room.capacity)
    
    pagination = query.order_by(Room.building, Room.room_number).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    rooms = pagination.items
    
    buildings = db.session.query(Room.building).distinct().all()
    buildings = [b[0] for b in buildings]
    
    return render_template('rooms/index.html',
                         title='房间管理',
                         rooms=rooms,
                         pagination=pagination,
                         buildings=buildings,
                         building_filter=building_filter,
                         status_filter=status_filter)


@bp.route('/add', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def add():
    """添加房间"""
    form = RoomForm()
    
    fee_standards = FeeStandard.query.filter_by(is_active=True).all()
    form.fee_standard_id.choices = [(0, '未选择')] + [(f.id, f'{f.name} ({f.price}/{f.unit})') for f in fee_standards]
    
    if form.validate_on_submit():
        existing = Room.query.filter_by(
            building=form.building.data,
            room_number=form.room_number.data
        ).first()
        
        if existing:
            flash(f'房间 {form.building}-{form.room_number} 已存在', 'danger')
            return redirect(url_for('rooms.add'))
        
        room = Room()
        form.populate_obj(room)
        
        if room.fee_standard_id == 0:
            room.fee_standard_id = None
        
        room.status = 'available'
        db.session.add(room)
        db.session.commit()
        
        flash(f'房间 {room.building}-{room.room_number} 添加成功！', 'success')
        return redirect(url_for('rooms.index'))
    
    return render_template('rooms/add.html', title='添加房间', form=form)


@bp.route('/edit/<int:room_id>', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def edit(room_id):
    """编辑房间"""
    room = Room.query.get_or_404(room_id)
    form = RoomForm(obj=room)
    
    fee_standards = FeeStandard.query.filter_by(is_active=True).all()
    form.fee_standard_id.choices = [(0, '未选择')] + [(f.id, f'{f.name} ({f.price}/{f.unit})') for f in fee_standards]
    
    if form.validate_on_submit():
        existing = Room.query.filter(
            Room.building == form.building.data,
            Room.room_number == form.room_number.data,
            Room.id != room_id
        ).first()
        
        if existing:
            flash(f'房间 {form.building.data}-{form.room_number.data} 已存在', 'danger')
            return redirect(url_for('rooms.edit', room_id=room_id))
        
        old_capacity = room.capacity
        
        form.populate_obj(room)
        
        if room.fee_standard_id == 0:
            room.fee_standard_id = None
        
        if old_capacity != room.capacity:
            room.current_occupancy = min(room.current_occupancy, room.capacity)
            if room.current_occupancy >= room.capacity:
                room.status = 'full'
            else:
                room.status = 'available'
        
        db.session.commit()
        flash(f'房间 {room.building}-{room.room_number} 信息已更新！', 'success')
        return redirect(url_for('rooms.index'))
    
    return render_template('rooms/edit.html', title='编辑房间', form=form, room=room)


@bp.route('/delete/<int:room_id>', methods=['POST'])
@login_required
@permission_required('write')
def delete(room_id):
    """删除房间"""
    room = Room.query.get_or_404(room_id)
    
    if room.current_occupancy > 0:
        flash(f'房间 {room.building}-{room.room_number} 仍有学生入住，无法删除', 'danger')
        return redirect(url_for('rooms.index'))
    
    db.session.delete(room)
    db.session.commit()
    
    flash(f'房间 {room.building}-{room.room_number} 已删除', 'success')
    return redirect(url_for('rooms.index'))


@bp.route('/detail/<int:room_id>')
@login_required
def detail(room_id):
    """房间详情"""
    room = Room.query.get_or_404(room_id)
    students = Student.query.filter_by(room_id=room_id, status='active').all()
    
    return render_template('rooms/detail.html', title=f'{room.building}-{room.room_number}',
                         room=room, students=students)


@bp.route('/batch-add', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def batch_add():
    """批量添加房间"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择要上传的文件', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('请选择要上传的文件', 'danger')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('只支持 Excel 文件 (.xlsx, .xls)', 'danger')
            return redirect(request.url)
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1] if cell.value]
            
            success_count = 0
            skip_count = 0
            error_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                row_data = dict(zip(headers, row))
                
                try:
                    building = str(row_data.get('楼号', '')).strip()
                    room_number = str(row_data.get('房间号', '')).strip()
                    capacity = int(row_data.get('容量', 2))
                    floor = row_data.get('楼层')
                    
                    if not building or not room_number:
                        error_count += 1
                        continue
                    
                    existing = Room.query.filter_by(building=building, room_number=room_number).first()
                    if existing:
                        skip_count += 1
                        continue
                    
                    room = Room(
                        building=building,
                        room_number=room_number,
                        capacity=capacity,
                        floor=floor,
                        status='available'
                    )
                    
                    fee_name = str(row_data.get('收费标准', '')).strip()
                    if fee_name:
                        fee_std = FeeStandard.query.filter(
                            (FeeStandard.name == fee_name) |
                            (FeeStandard.name.contains(fee_name))
                        ).first()
                        if fee_std:
                            room.fee_standard_id = fee_std.id
                    
                    db.session.add(room)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    continue
            
            db.session.commit()
            flash(f'导入完成！成功: {success_count}, 跳过(已存在): {skip_count}, 失败: {error_count}', 'success')
            return redirect(url_for('rooms.index'))
            
        except Exception as e:
            flash(f'导入失败: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('rooms/batch_add.html', title='批量添加房间')


@bp.route('/batch-edit', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def batch_edit():
    """批量修改房间"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择要上传的文件', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('请选择要上传的文件', 'danger')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('只支持 Excel 文件 (.xlsx, .xls)', 'danger')
            return redirect(request.url)
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1] if cell.value]
            
            success_count = 0
            error_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                row_data = dict(zip(headers, row))
                
                try:
                    building = str(row_data.get('楼号', '')).strip()
                    room_number = str(row_data.get('房间号', '')).strip()
                    
                    room = Room.query.filter_by(building=building, room_number=room_number).first()
                    if not room:
                        error_count += 1
                        continue
                    
                    if row_data.get('新房间号'):
                        room.room_number = str(row_data['新房间号']).strip()
                    if row_data.get('容量'):
                        room.capacity = int(row_data['容量'])
                    if row_data.get('楼层'):
                        room.floor = int(row_data['楼层'])
                    if row_data.get('备注'):
                        room.description = str(row_data['备注'])
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    continue
            
            db.session.commit()
            flash(f'修改完成！成功: {success_count}, 失败: {error_count}', 'success')
            return redirect(url_for('rooms.index'))
            
        except Exception as e:
            flash(f'修改失败: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('rooms/batch_edit.html', title='批量修改房间')


@bp.route('/status')
@login_required
def status():
    """房间状态统计"""
    total_rooms = Room.query.count()
    total_capacity = db.session.query(db.func.sum(Room.capacity)).scalar() or 0
    total_occupancy = db.session.query(db.func.sum(Room.current_occupancy)).scalar() or 0
    
    available_rooms = Room.query.filter(Room.current_occupancy == 0).count()  # 完全空的房间（无人入住）
    full_rooms = Room.query.filter(Room.current_occupancy >= Room.capacity).count()
    
    buildings = db.session.query(
        Room.building,
        db.func.count(Room.id).label('room_count'),
        db.func.sum(Room.capacity).label('total_capacity'),
        db.func.sum(Room.current_occupancy).label('current_occupancy')
    ).group_by(Room.building).all()
    
    return render_template('rooms/status.html',
                         title='房间状态',
                         total_rooms=total_rooms,
                         total_capacity=total_capacity,
                         total_occupancy=total_occupancy,
                         available_rooms=available_rooms,
                         full_rooms=full_rooms,
                         buildings=buildings)


@bp.route('/export-template')
@login_required
def export_template():
    """下载房间导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = '房间导入模板'
    
    # 表头
    headers = ['楼号', '房间号', '容量', '楼层', '收费标准']
    
    # 设置表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 15
    
    # 添加示例数据
    sample_data = [
        ['1号楼', '101', 2, 1, '标准双人间'],
        ['1号楼', '102', 2, 1, '标准双人间'],
        ['2号楼', '201', 4, 2, '四人间'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='房间导入模板.xlsx'
    )

@bp.route('/overview')
@login_required
def overview():
    """楼栋房间入住一览表 - 自动跳转到第一个楼栋"""
    # 获取第一个楼栋
    first_building = db.session.query(Room.building).distinct().order_by(Room.building).first()
    if first_building:
        return redirect(url_for('rooms.building_overview', building=first_building[0]))
    else:
        flash('暂无楼栋数据', 'warning')
        return redirect(url_for('rooms.index'))


def _build_room_info(room):
    """构建单个房间的入住信息（供平面图和表格视图共用）"""
    active_students = Student.query.filter_by(
        room_id=room.id, status='active'
    ).all()
    total_beds = room.capacity
    occupied_beds = sum(s.bed_occupancy for s in active_students)
    empty_beds = total_beds - occupied_beds
    if active_students:
        room_type = '单' if any(s.bed_occupancy == 2 for s in active_students) else '双'
    else:
        room_type = '双'
    return {
        'room': room,
        'students': active_students,
        'empty_beds': empty_beds,
        'room_type': room_type,
        'is_empty': len(active_students) == 0,
        'has_partial': len(active_students) > 0 and empty_beds > 0
    }


@bp.route('/building/<building>/overview')
@login_required
def building_overview(building):
    """楼栋房间入住情况一览表（支持平面图和表格两种视图）"""
    import re as _re
    view = request.args.get('view', 'plan')  # plan=平面图(默认), table=表格

    rooms = Room.query.filter_by(building=building).order_by(Room.floor, Room.room_number).all()

    if not rooms:
        flash(f'楼栋 {building} 不存在或暂无房间', 'warning')
        return redirect(url_for('rooms.index'))

    buildings = [b[0] for b in db.session.query(Room.building).distinct().order_by(Room.building).all()]

    # 按楼层分组
    floors_data = {}
    room_info_map = {}
    for room in rooms:
        floor = room.floor or 0
        info = _build_room_info(room)
        room_info_map[room.room_number] = info
        floors_data.setdefault(floor, []).append(info)

    total_rooms = len(rooms)
    empty_rooms = sum(1 for f in floors_data.values() for r in f if r['is_empty'])
    partial_rooms = sum(1 for f in floors_data.values() for r in f if r['has_partial'])
    full_rooms = total_rooms - empty_rooms - partial_rooms
    total_empty_beds = sum(r['empty_beds'] for f in floors_data.values() for r in f)
    
    # 统计入住人数和性别分布
    total_students = 0
    male_count = 0
    female_count = 0
    for f_data in floors_data.values():
        for r in f_data:
            for s in r['students']:
                total_students += 1
                if s.gender and '男' in s.gender:
                    male_count += 1
                elif s.gender and '女' in s.gender:
                    female_count += 1

    floor_stats = {}
    for floor, f_data in floors_data.items():
        floor_students = 0
        floor_male = 0
        floor_female = 0
        for r in f_data:
            for s in r['students']:
                floor_students += 1
                if s.gender and '男' in s.gender:
                    floor_male += 1
                elif s.gender and '女' in s.gender:
                    floor_female += 1
        floor_stats[floor] = {
            'total': len(f_data),
            'empty_rooms': sum(1 for r in f_data if r['is_empty']),
            'empty_beds': sum(r['empty_beds'] for r in f_data),
            'students': floor_students,
            'male': floor_male,
            'female': floor_female
        }

    floors_sorted = sorted(floors_data.keys(), reverse=True)

    stats = {
        'total': total_rooms, 'empty': empty_rooms,
        'partial': partial_rooms, 'full': full_rooms,
        'empty_beds': total_empty_beds,
        'total_students': total_students,
        'male_count': male_count,
        'female_count': female_count
    }

    # 判断是否为A塔/B塔（有平面图配置的楼栋）
    tower = detect_tower(rooms)
    layout = get_layout(tower) if tower else None

    # 构建平面图数据：每层一个网格
    floor_plans = {}
    if layout:
        for floor in floors_sorted:
            floor_str = str(floor)
            cells_render = []
            for cell in layout['cells']:
                c = dict(cell)
                if cell['type'] == 'room':
                    # 房间号 = 塔字母 + 楼层号(2位) + 后缀
                    full_num = f"{tower}{floor_str.zfill(2)}{cell['suffix']}"
                    c['room_num'] = full_num
                    info = room_info_map.get(full_num)
                    if info:
                        c['room_id'] = info['room'].id
                        c['is_empty'] = info['is_empty']
                        c['has_partial'] = info['has_partial']
                        c['is_full'] = not info['is_empty'] and not info['has_partial']
                        c['room_type'] = info['room_type']
                        c['empty_beds'] = info['empty_beds']
                        c['students'] = info['students']
                        c['exists'] = True
                    else:
                        # 布局中有但系统未录入 = 预留房
                        c['exists'] = False
                        c['is_empty'] = True
                        c['has_partial'] = False
                        c['is_full'] = False
                cells_render.append(c)
            floor_plans[floor] = {
                'rows': layout['rows'],
                'cols': layout['cols'],
                'cells': cells_render
            }

    # 如果没有平面图配置，强制使用表格视图
    has_plan = layout is not None
    if not has_plan:
        view = 'table'

    return render_template('rooms/building_overview.html',
                         title=f'{building}房间一览',
                         building=building,
                         view=view,
                         has_plan=has_plan,
                         tower=tower,
                         floors_data=floors_data,
                         floors_sorted=floors_sorted,
                         floor_stats=floor_stats,
                         floor_plans=floor_plans,
                         buildings=buildings,
                         stats=stats)


@bp.route('/<int:room_id>/toggle-type', methods=['POST'])
@login_required
@permission_required('write')
def toggle_type(room_id):
    """切换房间户型（单人间/双人间）- 同时更新该房间内所有学生的收费标准"""
    room = Room.query.get_or_404(room_id)
    
    active_students = Student.query.filter_by(room_id=room_id, status='active').all()
    
    if not active_students:
        flash('该房间没有入住学生，无法切换户型', 'warning')
        return redirect(url_for('rooms.detail', room_id=room_id))
    
    # 判断当前户型：如果有任何学生bed_occupancy=2，当前为单人间，否则为双人间
    is_current_single = any(s.bed_occupancy == 2 for s in active_students)
    
    if is_current_single:
        # 单人间 → 双人间：所有学生bed_occupancy改为1，房间释放床位
        for s in active_students:
            s.bed_occupancy = 1
        room.current_occupancy = len(active_students)  # 每人占1个床位
        # 更新收费标准为双人间（找到含"双人间"的收费标准）
        double_fee = FeeStandard.query.filter(FeeStandard.name.contains('双人间'), FeeStandard.is_active == True).first()
        if double_fee:
            for s in active_students:
                s.fee_standard_id = double_fee.id
            room.fee_standard_id = double_fee.id
        flash(f'房间 {room.building}-{room.room_number} 已从单人间切换为双人间，释放1个床位', 'success')
    else:
        # 双人间 → 单人间：只能有一个学生，bed_occupancy改为2
        if len(active_students) > 1:
            flash(f'该房间有 {len(active_students)} 名学生，无法切换为单人间（单人间只能住1人）', 'danger')
            return redirect(url_for('rooms.detail', room_id=room_id))
        
        for s in active_students:
            s.bed_occupancy = 2
        room.current_occupancy = 2  # 单人间占2个床位
        # 更新收费标准为单人间（找到含"单人间"的收费标准）
        single_fee = FeeStandard.query.filter(FeeStandard.name.contains('单人间'), FeeStandard.is_active == True).first()
        if single_fee:
            for s in active_students:
                s.fee_standard_id = single_fee.id
            room.fee_standard_id = single_fee.id
        flash(f'房间 {room.building}-{room.room_number} 已从双人间切换为单人间', 'success')
    
    # 更新房间状态
    if room.current_occupancy >= room.capacity:
        room.status = 'full'
    else:
        room.status = 'available'
    
    db.session.commit()
    return redirect(url_for('rooms.detail', room_id=room_id))
