# -*- coding: utf-8 -*-
"""
入住计划管理路由模块
提供入住计划相关功能
"""
from flask import render_template, Blueprint, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required
from flask_wtf import FlaskForm
from wtforms import StringField, DateField, IntegerField
from wtforms.validators import DataRequired, Optional, NumberRange
from app.models import db
from app.models import Reservation, Room, Student
from app.decorators import permission_required
from datetime import datetime, date, timedelta
from calendar import monthrange
import io

bp = Blueprint('reservations', __name__, url_prefix='/reservations')


class ReservationForm(FlaskForm):
    """入住计划表单"""
    department = StringField('部门', validators=[Optional()])
    group_name = StringField('团体名称', validators=[Optional()])
    person_count = IntegerField('入住人数', validators=[DataRequired(message='请输入入住人数'), NumberRange(min=1, max=200)])
    check_in_date = DateField('入住时间', format='%Y-%m-%d', validators=[DataRequired(message='请选择入住时间')])
    check_out_date = DateField('离开时间', format='%Y-%m-%d', validators=[Optional()])
    notes = StringField('备注', validators=[Optional()])


@bp.route('/')
@login_required
def index():
    """入住计划列表"""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    status_filter = request.args.get('status', '')
    month_filter = request.args.get('month', '')
    
    query = Reservation.query.filter(Reservation.status != 'cancelled')
    
    if status_filter:
        query = query.filter(Reservation.status == status_filter)
    
    if month_filter:
        try:
            year, month = map(int, month_filter.split('-'))
            start = date(year, month, 1)
            end = date(year, month, monthrange(year, month)[1])
            query = query.filter(
                Reservation.check_in_date <= end,
                db.or_(
                    Reservation.check_out_date >= start,
                    Reservation.check_out_date == None
                )
            )
        except:
            pass
    
    pagination = query.order_by(Reservation.check_in_date).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    reservations = pagination.items
    
    return render_template('reservations/index.html', title='入住计划',
                         reservations=reservations, pagination=pagination,
                         status_filter=status_filter, month_filter=month_filter)


@bp.route('/add', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def add():
    """添加入住计划"""
    form = ReservationForm()
    
    if form.validate_on_submit():
        reservation = Reservation()
        form.populate_obj(reservation)
        
        # 自动计算需要房间数（每间2人）
        total_beds = 2  # 所有房间都是2张床
        reservation.rooms_needed = (reservation.person_count + total_beds - 1) // total_beds
        
        if reservation.check_out_date and reservation.check_in_date:
            reservation.duration_days = (reservation.check_out_date - reservation.check_in_date).days
        
        reservation.status = 'pending'
        db.session.add(reservation)
        db.session.commit()
        
        flash(f'入住计划已添加！需要 {reservation.rooms_needed} 间房间', 'success')
        return redirect(url_for('reservations.index'))
    
    return render_template('reservations/add.html', title='添加入住计划', form=form)


@bp.route('/edit/<int:reservation_id>', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def edit(reservation_id):
    """编辑入住计划"""
    reservation = Reservation.query.get_or_404(reservation_id)
    form = ReservationForm(obj=reservation)
    
    if form.validate_on_submit():
        form.populate_obj(reservation)
        
        # 重新计算需要房间数
        total_beds = 2
        reservation.rooms_needed = (reservation.person_count + total_beds - 1) // total_beds
        
        if reservation.check_out_date and reservation.check_in_date:
            reservation.duration_days = (reservation.check_out_date - reservation.check_in_date).days
        
        db.session.commit()
        
        flash('入住计划已更新！', 'success')
        return redirect(url_for('reservations.index'))
    
    return render_template('reservations/edit.html', title='编辑入住计划', form=form, reservation=reservation)


@bp.route('/delete/<int:reservation_id>', methods=['POST'])
@login_required
@permission_required('write')
def delete(reservation_id):
    """删除入住计划"""
    reservation = Reservation.query.get_or_404(reservation_id)
    
    db.session.delete(reservation)
    db.session.commit()
    
    flash('入住计划已删除', 'success')
    return redirect(url_for('reservations.index'))


@bp.route('/calendar')
@login_required
def calendar():
    """房间日历视图 - 显示某月每天的房间占用情况
    
    同时考虑入住计划(Reservation)和实际入住学生(Student)的房间占用
    """
    # 获取年月参数，默认当前月
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', date.today().month, type=int)
    
    # 确保月份在1-12之间
    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1
    
    # 获取总房间数
    total_rooms = Room.query.count()
    
    # 计算该月天数
    _, days_in_month = monthrange(year, month)
    month_start = date(year, month, 1)
    month_end = date(year, month, days_in_month)
    
    # 获取该月所有入住计划
    reservations = Reservation.query.filter(
        Reservation.check_in_date <= month_end,
        db.or_(
            Reservation.check_out_date >= month_start,
            Reservation.check_out_date == None
        ),
        Reservation.status != 'cancelled'
    ).all()
    
    # 获取该月所有已入住的学生（根据check_in_date和预计离开日期）
    housed_students = Student.query.filter(
        Student.status == 'active',
        Student.room_id != None,
        Student.check_in_date <= month_end,
        db.or_(
            Student.check_out_date >= month_start,
            Student.check_out_date == None
        )
    ).all()
    
    # 统计实际入住的学生房间占用（按床位占用数计算）
    # 注意：单人间的bed_occupancy=2，双人间bed_occupancy=1
    student_room_occupancy = {}  # {room_id: occupied_beds}
    for student in housed_students:
        if student.room_id:
            if student.room_id not in student_room_occupancy:
                student_room_occupancy[student.room_id] = 0
            student_room_occupancy[student.room_id] += student.bed_occupancy
    
    # 获取所有房间信息用于计算实际占用的房间数
    all_rooms = {r.id: r for r in Room.query.all()}
    
    # 计算实际入住学生占用的房间数
    actual_occupied_rooms = 0
    for room_id, occupied_beds in student_room_occupancy.items():
        room = all_rooms.get(room_id)
        if room and occupied_beds >= room.capacity:
            # 如果占用的床位数达到房间容量，则算占用1间
            actual_occupied_rooms += 1
        elif room:
            # 部分占用也算占用1间（因为房间已被使用）
            actual_occupied_rooms += 1
    
    # 生成每天的数据
    calendar_days = []
    for day in range(1, days_in_month + 1):
        current_date = date(year, month, day)
        
        # 计算当天入住计划占用的房间数
        occupied_rooms_from_reservations = 0
        day_reservations = []
        
        for res in reservations:
            # 检查计划是否在当天有效
            if res.check_in_date <= current_date:
                if res.check_out_date is None or res.check_out_date > current_date:
                    occupied_rooms_from_reservations += res.rooms_needed
                    day_reservations.append(res)
        
        # 计算当天实际入住学生占用的房间数
        # 获取当前日期在入住时间段内的学生
        day_students = []
        for student in housed_students:
            if student.check_in_date <= current_date:
                if student.check_out_date is None or student.check_out_date > current_date:
                    day_students.append(student)
        
        # 计算这些学生占用的房间数
        day_student_occupancy = {}
        for student in day_students:
            if student.room_id:
                if student.room_id not in day_student_occupancy:
                    day_student_occupancy[student.room_id] = 0
                day_student_occupancy[student.room_id] += student.bed_occupancy
        
        actual_rooms_for_day = 0
        for room_id, occupied_beds in day_student_occupancy.items():
            room = all_rooms.get(room_id)
            if room and occupied_beds >= room.capacity:
                actual_rooms_for_day += 1
            elif room:
                actual_rooms_for_day += 1
        
        # 总占用 = 计划占用 + 实际入住占用（取较大值，避免重复计算）
        # 如果有实际入住，应该反映真实占用情况
        occupied_rooms = max(occupied_rooms_from_reservations, actual_rooms_for_day)
        
        # 计算剩余房间数
        available_rooms = total_rooms - occupied_rooms
        
        # 判断是否缺少房间
        shortage = 0
        if occupied_rooms > total_rooms:
            shortage = occupied_rooms - total_rooms
        
        calendar_days.append({
            'date': current_date,
            'day': day,
            'occupied_rooms': occupied_rooms,
            'occupied_from_reservations': occupied_rooms_from_reservations,
            'occupied_from_students': actual_rooms_for_day,
            'available_rooms': max(0, available_rooms),
            'shortage': shortage,
            'total_rooms': total_rooms,
            'is_weekend': current_date.weekday() >= 5,
            'is_today': current_date == date.today(),
            'reservations': day_reservations,
            'students': day_students
        })
    
    # 获取高峰期预警（缺房的日子）
    peak_days = [d for d in calendar_days if d['shortage'] > 0]
    
    # 统计信息
    stats = {
        'total_rooms': total_rooms,
        'month_reservations': len(reservations),
        'total_persons': sum(r.person_count for r in reservations),
        'housed_students': len(housed_students),
        'actual_occupied_rooms': actual_occupied_rooms,
        'peak_days_count': len(peak_days)
    }
    
    return render_template('reservations/calendar.html', title='房间日历',
                         calendar_days=calendar_days,
                         year=year, month=month,
                         stats=stats, peak_days=peak_days,
                         today=date.today())


@bp.route('/api/daily-rooms')
@login_required
def api_daily_rooms():
    """API: 获取指定日期的房间占用详情"""
    target_date = request.args.get('date')
    if not target_date:
        return jsonify({'error': '缺少日期参数'}), 400
    
    try:
        target = datetime.strptime(target_date, '%Y-%m-%d').date()
    except:
        return jsonify({'error': '日期格式错误'}), 400
    
    total_rooms = Room.query.count()
    
    # 获取当天有效的入住计划
    reservations = Reservation.query.filter(
        Reservation.check_in_date <= target,
        db.or_(
            Reservation.check_out_date > target,
            Reservation.check_out_date == None
        ),
        Reservation.status != 'cancelled'
    ).all()
    
    occupied = sum(r.rooms_needed for r in reservations)
    
    return jsonify({
        'date': target.isoformat(),
        'total_rooms': total_rooms,
        'occupied_rooms': occupied,
        'available_rooms': max(0, total_rooms - occupied),
        'shortage': max(0, occupied - total_rooms),
        'reservations': [{
            'id': r.id,
            'department': r.department or '-',
            'group_name': r.group_name or r.student_name or '-',
            'person_count': r.person_count,
            'rooms_needed': r.rooms_needed,
            'check_in': r.check_in_date.isoformat() if r.check_in_date else '-',
            'check_out': r.check_out_date.isoformat() if r.check_out_date else '-',
            'status': r.status,
            'notes': r.notes or ''
        } for r in reservations]
    })


@bp.route('/stats')
@login_required
def stats():
    """房间统计面板"""
    total_rooms = Room.query.count()
    # 当前可用 = 完全空的房间（无人入住）
    available_rooms = Room.query.filter(Room.current_occupancy == 0).count()
    
    # 获取所有有效入住计划
    all_reservations = Reservation.query.filter(
        Reservation.status != 'cancelled'
    ).all()
    
    # 获取未来90天的数据
    future_date = date.today() + timedelta(days=90)
    future_reservations = [r for r in all_reservations 
                          if r.check_in_date <= future_date 
                          and (r.check_out_date is None or r.check_out_date >= date.today())]
    
    # 按月统计
    monthly_stats = {}
    for res in future_reservations:
        if res.check_in_date:
            month_key = res.check_in_date.strftime('%Y-%m')
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {'plans': 0, 'persons': 0, 'rooms': 0}
            monthly_stats[month_key]['plans'] += 1
            monthly_stats[month_key]['persons'] += res.person_count
            monthly_stats[month_key]['rooms'] += res.rooms_needed
    
    # 找出高峰期
    peak_analysis = []
    for d in range(0, 91):
        current = date.today() + timedelta(days=d)
        occupied = 0
        for res in all_reservations:
            if res.check_in_date <= current:
                if res.check_out_date is None or res.check_out_date > current:
                    if res.check_in_date <= future_date:
                        occupied += res.rooms_needed
        
        if occupied > total_rooms:
            peak_analysis.append({
                'date': current,
                'occupied': occupied,
                'shortage': occupied - total_rooms
            })
    
    return render_template('reservations/stats.html', title='房间统计',
                         total_rooms=total_rooms,
                         available_rooms=available_rooms,
                         monthly_stats=monthly_stats,
                         peak_analysis=peak_analysis[:10])


@bp.route('/batch-import', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def batch_import():
    """批量导入入住计划 - 支持用户模板格式"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择要上传的文件', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('请选择要上传的文件', 'danger')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('只支持 Excel 文件 (.xlsx, .xls)', 'danger')
            return redirect(request.url)
        
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.datetime import from_excel
            
            wb = load_workbook(file)
            
            # 优先读取"总表"工作表，否则读取活动工作表
            if '总表' in wb.sheetnames:
                ws = wb['总表']
            else:
                ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in ws[1] if cell.value]
            
            success_count = 0
            error_count = 0
            error_details = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0] and not any(row[:5]):
                    continue
                
                try:
                    # 构建行数据字典
                    row_data = dict(zip(headers, row))
                    
                    # 提取数据
                    department = str(row_data.get('部门', '')).strip()
                    group_name = str(row_data.get('国籍/团体名称', '') or row_data.get('团体名称', '')).strip()
                    person_count_str = str(row_data.get('入住人数', 0)).strip()
                    
                    # 处理Excel日期序列号
                    check_in_date = row_data.get('入住时间', '')
                    check_out_date = row_data.get('离开时间', '')
                    
                    if isinstance(check_in_date, (int, float)):
                        check_in_date = from_excel(check_in_date)
                        if hasattr(check_in_date, 'date'):
                            check_in_date = check_in_date.date()
                    
                    if isinstance(check_out_date, (int, float)):
                        check_out_date = from_excel(check_out_date)
                        if hasattr(check_out_date, 'date'):
                            check_out_date = check_out_date.date()
                    
                    if not check_in_date:
                        error_count += 1
                        error_details.append(f'第{row_idx}行: 缺少入住时间')
                        continue
                    
                    # 解析日期
                    if isinstance(check_in_date, date):
                        pass
                    else:
                        check_in_date = datetime.strptime(str(check_in_date), '%Y-%m-%d').date()
                    
                    if check_out_date:
                        if isinstance(check_out_date, date):
                            pass
                        else:
                            check_out_date = datetime.strptime(str(check_out_date), '%Y-%m-%d').date()
                    
                    # 解析人数
                    try:
                        person_count = int(person_count_str) if person_count_str else 0
                    except:
                        person_count = 0
                    
                    # 解析单人间数量
                    single_rooms_str = str(row_data.get('单人间数量', 0) or row_data.get('单人间', 0) or 0).strip()
                    try:
                        single_rooms = int(single_rooms_str) if single_rooms_str else 0
                    except:
                        single_rooms = 0
                    
                    # 解析双人间数量
                    double_rooms_str = str(row_data.get('双人间数量', 0) or row_data.get('双人间', 0) or 0).strip()
                    try:
                        double_rooms = int(double_rooms_str) if double_rooms_str else 0
                    except:
                        double_rooms = 0
                    
                    # 解析用户指定的房间数
                    rooms_str = str(row_data.get('需要房间数', 0)).strip()
                    try:
                        rooms_needed_user = int(rooms_str) if rooms_str else 0
                    except:
                        rooms_needed_user = 0
                    
                    # 计算房间数：优先使用用户指定的值，否则根据单人间/双人间计算
                    if rooms_needed_user > 0:
                        # 用户明确指定了房间数，直接使用
                        rooms_needed = rooms_needed_user
                        # 如果入住人数为0，根据房间数估算（假设双人间）
                        if person_count <= 0:
                            person_count = rooms_needed * 2
                    elif single_rooms > 0 or double_rooms > 0:
                        # 根据单人间和双人间计算
                        rooms_needed = single_rooms + double_rooms
                        # 单人间1人，双人间2人
                        person_count = single_rooms * 1 + double_rooms * 2 if person_count <= 0 else person_count
                    else:
                        # 都没填，按入住人数自动计算（默认双人间）
                        if person_count <= 0:
                            person_count = 2
                        # 向上取整
                        rooms_needed = (person_count + 1) // 2
                    
                    # 创建入住计划
                    reservation = Reservation(
                        department=department if department and department != 'None' else '',
                        group_name=group_name if group_name and group_name != 'None' else '',
                        person_count=person_count,
                        rooms_needed=rooms_needed,
                        student_name='',  # 数据库约束要求非空，给空字符串
                        check_in_date=check_in_date,
                        check_out_date=check_out_date,
                        status='pending',
                        notes=str(row_data.get('备注', '')).strip() if row_data.get('备注') else ''
                    )
                    
                    # 计算天数
                    if reservation.check_out_date:
                        reservation.duration_days = (reservation.check_out_date - reservation.check_in_date).days
                    
                    db.session.add(reservation)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_details.append(f'第{row_idx}行: {str(e)}')
                    continue
            
            db.session.commit()
            
            msg = f'导入完成！成功: {success_count} 条'
            if error_count > 0:
                msg += f'，失败: {error_count} 条'
            flash(msg, 'success' if error_count == 0 else 'warning')
            
            # 存储错误详情到session供查看
            if error_details:
                from flask import session
                session['import_errors'] = error_details[:20]
            
            return redirect(url_for('reservations.index'))
            
        except Exception as e:
            flash(f'导入失败: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('reservations/batch_import.html', title='批量导入入住计划')


@bp.route('/template-download')
@login_required
def template_download():
    """下载导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = '总表'
    
    # 设置表头样式
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    headers = ['部门', '国籍/团体名称', '入住时间', '离开时间', '入住人数', '单人间数量', '双人间数量', '需要房间数', '备注']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加示例数据
    example_data = [
        ['国际交流处', '美国交换生团', '2024-03-01', '2024-06-30', 20, 0, 10, 10, '春季学期交换生'],
        ['人事处', '新入职教师', '2024-03-05', '2024-08-31', 6, 2, 2, 4, '2人需单人间'],
        ['国际交流处', '德国暑期研学团', '2024-07-01', '2024-07-31', 40, 0, 20, 20, ''],
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    
    # 保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='入住计划导入模板.xlsx'
    )


@bp.route('/confirm/<int:reservation_id>', methods=['POST'])
@login_required
@permission_required('write')
def confirm(reservation_id):
    """确认入住计划"""
    reservation = Reservation.query.get_or_404(reservation_id)
    
    # 检查是否有足够的空房间
    total_rooms = Room.query.count()
    target_date = reservation.check_in_date
    
    # 计算当天已占用房间数
    existing_reservations = Reservation.query.filter(
        Reservation.id != reservation_id,
        Reservation.check_in_date <= target_date,
        db.or_(
            Reservation.check_out_date > target_date,
            Reservation.check_out_date == None
        ),
        Reservation.status == 'confirmed'
    ).all()
    
    occupied = sum(r.rooms_needed for r in existing_reservations)
    
    if occupied + reservation.rooms_needed > total_rooms:
        flash(f'房间不足！需要 {reservation.rooms_needed} 间，当前剩余 {total_rooms - occupied} 间', 'danger')
        return redirect(url_for('reservations.index'))
    
    reservation.status = 'confirmed'
    db.session.commit()
    
    flash(f'入住计划已确认！{reservation.group_name or reservation.student_name} ({reservation.person_count}人, {reservation.rooms_needed}间)', 'success')
    return redirect(url_for('reservations.index'))


@bp.route('/cancel/<int:reservation_id>', methods=['POST'])
@login_required
@permission_required('write')
def cancel(reservation_id):
    """取消入住计划"""
    reservation = Reservation.query.get_or_404(reservation_id)
    reservation.status = 'cancelled'
    db.session.commit()
    
    flash('入住计划已取消', 'info')
    return redirect(url_for('reservations.index'))
