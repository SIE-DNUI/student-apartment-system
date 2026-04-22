# -*- coding: utf-8 -*-
"""
学生管理路由模块
提供学生信息管理功能
"""
from flask import render_template, Blueprint, redirect, url_for, flash, request, send_file, session
from flask_login import login_required
from flask_wtf import FlaskForm
from wtforms import StringField, SelectField, DateField, TextAreaField
from wtforms.validators import DataRequired, Optional
from wtforms.widgets import TextArea
from app.models import db
from app.models import Student, Room, FeeStandard, FeeRecord, Alert
from app.decorators import permission_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, date, timedelta
import io

bp = Blueprint('students', __name__, url_prefix='/students')


class StudentForm(FlaskForm):
    """学生表单"""
    name = StringField('姓名', validators=[DataRequired(message='请输入姓名')])
    gender = SelectField('性别', choices=[('男', '男'), ('女', '女'), ('其他', '其他')], validators=[Optional()])
    nationality = StringField('国籍', validators=[Optional()])
    department = StringField('所属业务部', validators=[Optional()])
    major = StringField('专业', validators=[Optional()])
    room_id = SelectField('分配房间', coerce=int, validators=[Optional()])
    check_in_date = DateField('入住日期', format='%Y-%m-%d', validators=[Optional()])
    check_out_date = DateField('预计离开日期', format='%Y-%m-%d', validators=[Optional()])
    fee_standard_id = SelectField('收费标准', coerce=int, validators=[Optional()])
    payment_due_date = DateField('缴费到期日期', format='%Y-%m-%d', validators=[Optional()])
    residence_permit_expiry = DateField('居留许可到期时间', format='%Y-%m-%d', validators=[Optional()])
    notes = StringField('备注', widget=TextArea(), validators=[Optional()])


@bp.route('/')
@login_required
def index():
    """学生列表"""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    search = request.args.get('search', '')
    major = request.args.get('major', '')
    filter_status = request.args.get('filter', 'all')  # all, housed, unhoused
    
    # 排除已归档的学生
    query = Student.query.filter(Student.status != 'archived')
    
    if search:
        # 搜索姓名、性别、国籍、房间号、所属业务部
        # 房间号需要关联Room表搜索
        from sqlalchemy import or_
        query = query.outerjoin(Room, Student.room_id == Room.id)
        query = query.filter(
            or_(
                Student.name.contains(search),
                Student.gender.contains(search),
                Student.nationality.contains(search),
                Student.department.contains(search),
                Room.room_number.contains(search),
                Room.building.contains(search)  # 也支持搜索楼栋
            )
        )
    
    # 按专业筛选
    if major:
        query = query.filter(Student.major.contains(major))
    
    # 按住宿状态筛选
    if filter_status == 'housed':
        query = query.filter(Student.room_id != None)
    elif filter_status == 'unhoused':
        query = query.filter(Student.room_id == None)
    
    pagination = query.order_by(Student.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    students = pagination.items
    
    # 获取居留许可即将到期的学生数量（排除已归档）
    expiring_count = Student.query.filter(
        Student.status != 'archived',
        Student.residence_permit_expiry != None,
        Student.residence_permit_expiry <= date.today() + timedelta(days=30),
        Student.residence_permit_expiry >= date.today()
    ).count()
    
    # 计算欠费学生数量
    arrears_count = 0
    for s in students:
        if s.has_arrears():
            arrears_count += 1
    
    return render_template('students/index.html', 
                         title='学生管理',
                         students=students,
                         pagination=pagination,
                         search=search,
                         major=major,
                         filter_status=filter_status,
                         expiring_count=expiring_count,
                         arrears_count=arrears_count)


@bp.route('/add', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def add():
    """添加学生"""
    form = StudentForm()
    
    # 预填的房间ID（从URL参数获取）
    prefill_room_id = request.args.get('room_id', type=int)
    
    available_rooms = Room.query.filter(Room.current_occupancy < Room.capacity).all()
    form.room_id.choices = [(0, '未分配')] + [(r.id, f'{r.building}-{r.room_number}') for r in available_rooms]
    
    fee_standards = FeeStandard.query.filter_by(is_active=True).all()
    form.fee_standard_id.choices = [(0, '未选择')] + [(f.id, f'{f.name} ({f.price}/{f.unit})') for f in fee_standards]
    
    # 如果有预填的房间ID，设置默认值
    if prefill_room_id:
        room = Room.query.get(prefill_room_id)
        if room:
            # 将预填房间添加到可选列表中（即使已满）
            existing_choice = [(r.id, f'{r.building}-{r.room_number}') for r in available_rooms]
            if prefill_room_id not in [c[0] for c in existing_choice]:
                # 房间已满但用户仍想添加入住（如调换房间场景）
                form.room_id.choices = [(0, '未分配')] + [(r.id, f'{r.building}-{r.room_number}') for r in available_rooms] + [(room.id, f'{room.building}-{room.room_number} (已满)')]
            form.room_id.data = prefill_room_id
    
    if form.validate_on_submit():
        student = Student()
        form.populate_obj(student)
        
        if student.room_id == 0:
            student.room_id = None
        
        if student.fee_standard_id == 0:
            student.fee_standard_id = None
        
        # 处理床位占用数
        bed_occupancy = request.form.get('bed_occupancy', '1', type=int)
        student.bed_occupancy = bed_occupancy
        
        # 处理本次缴纳房费
        current_payment = request.form.get('current_payment', '0', type=float)
        student.total_paid = current_payment if current_payment else 0
        
        if student.room_id:
            room = Room.query.get(student.room_id)
            if room:
                # 根据床位占用数更新房间入住人数
                room.current_occupancy += bed_occupancy
                if room.current_occupancy >= room.capacity:
                    room.status = 'full'
        
        student.status = 'active'
        db.session.add(student)
        db.session.commit()
        
        flash('学生添加成功！', 'success')
        return redirect(url_for('students.index'))
    
    return render_template('students/add.html', title='添加学生', form=form)


@bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def edit(id):
    """编辑学生"""
    student = Student.query.get_or_404(id)
    form = StudentForm(obj=student)
    
    # 获取来源页码，用于返回原页面
    page = request.args.get('page', 1, type=int)
    
    available_rooms = Room.query.filter(
        (Room.current_occupancy < Room.capacity) | (Room.id == student.room_id)
    ).all()
    form.room_id.choices = [(0, '未分配')] + [(r.id, f'{r.building}-{r.room_number}') for r in available_rooms]
    
    fee_standards = FeeStandard.query.filter_by(is_active=True).all()
    form.fee_standard_id.choices = [(0, '未选择')] + [(f.id, f'{f.name} ({f.price}/{f.unit})') for f in fee_standards]
    
    if form.validate_on_submit():
        old_room_id = student.room_id
        
        form.populate_obj(student)
        
        if student.room_id == 0:
            student.room_id = None
        
        if student.fee_standard_id == 0:
            student.fee_standard_id = None
        
        if old_room_id != student.room_id:
            if old_room_id:
                old_room = Room.query.get(old_room_id)
                if old_room:
                    old_room.current_occupancy -= 1
                    old_room.status = 'available'
            
            if student.room_id:
                new_room = Room.query.get(student.room_id)
                if new_room:
                    new_room.current_occupancy += 1
                    if new_room.current_occupancy >= new_room.capacity:
                        new_room.status = 'full'
        
        db.session.commit()
        flash('学生信息更新成功！', 'success')
        return redirect(url_for('students.index', page=page))
    
    return render_template('students/edit.html', title='编辑学生', form=form, student=student, page=page)


@bp.route('/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('write')
def delete(id):
    """删除学生（归档，保留3年）"""
    student = Student.query.get_or_404(id)
    
    # 释放房间
    if student.room_id:
        room = Room.query.get(student.room_id)
        if room:
            room.current_occupancy -= 1
            if room.current_occupancy < room.capacity:
                room.status = 'available'
    
    # 归档学生信息（保存房间ID用于成本统计）
    student.status = 'archived'
    if student.room_id:
        student.archived_room_id = student.room_id  # 保存房间ID用于归档后的成本统计
    student.room_id = None
    student.deleted_at = datetime.utcnow()
    student.retention_until = date.today() + timedelta(days=365*3)  # 保留3年
    
    db.session.commit()
    flash(f'学生 {student.name} 已删除并归档，将保留至 {student.retention_until}', 'success')
    return redirect(url_for('students.index'))


@bp.route('/<int:id>/checkout', methods=['POST'])
@login_required
@permission_required('write')
def checkout(id):
    """单个学生退房"""
    student = Student.query.get_or_404(id)
    
    if not student.room_id:
        flash('该学生未入住，无需退房', 'warning')
        return redirect(url_for('students.index'))
    
    # 更新房间入住人数
    room = Room.query.get(student.room_id)
    if room:
        room.current_occupancy -= 1
        if room.current_occupancy < room.capacity:
            room.status = 'available'
    
    # 清空学生的房间信息并归档（保存房间ID用于成本统计）
    if student.room_id:
        student.archived_room_id = student.room_id  # 保存房间ID用于归档后的成本统计
    student.room_id = None
    student.check_out_date = date.today()
    student.status = 'archived'
    student.deleted_at = datetime.utcnow()
    student.retention_until = date.today() + timedelta(days=365*3)  # 保留3年
    
    db.session.commit()
    
    flash(f'学生 {student.name} 已退房并归档，将保留至 {student.retention_until}', 'success')
    return redirect(url_for('students.index'))


@bp.route('/batch-checkout', methods=['POST'])
@login_required
@permission_required('write')
def batch_checkout():
    """批量退房"""
    student_ids = request.form.getlist('student_ids')
    
    if not student_ids:
        flash('请选择要退房的学生', 'warning')
        return redirect(url_for('students.index'))
    
    count = 0
    for student_id in student_ids:
        student = Student.query.get(int(student_id))
        if student and student.room_id:
            # 更新房间入住人数
            room = Room.query.get(student.room_id)
            if room:
                room.current_occupancy -= 1
                if room.current_occupancy < room.capacity:
                    room.status = 'available'
            
            # 清空学生的房间信息并归档（保存房间ID用于成本统计）
            student.archived_room_id = student.room_id  # 保存房间ID用于归档后的成本统计
            student.room_id = None
            student.check_out_date = date.today()
            student.status = 'archived'
            student.deleted_at = datetime.utcnow()
            student.retention_until = date.today() + timedelta(days=365*3)  # 保留3年
            count += 1
    
    db.session.commit()
    
    flash(f'已成功退房 {count} 名学生', 'success')
    return redirect(url_for('students.index'))


@bp.route('/batch-edit', methods=['POST'])
@login_required
@permission_required('write')
def batch_edit():
    """批量修改学生信息"""
    student_ids = request.form.getlist('student_ids')
    
    if not student_ids:
        flash('请选择要修改的学生', 'warning')
        return redirect(url_for('students.index'))
    
    count = 0
    for student_id in student_ids:
        student = Student.query.get(int(student_id))
        if student:
            # 只修改填写了的字段
            major = request.form.get('major', '').strip()
            if major:
                student.major = major
            
            department = request.form.get('department', '').strip()
            if department:
                student.department = department
            
            check_in_date = request.form.get('check_in_date', '').strip()
            if check_in_date:
                try:
                    student.check_in_date = datetime.strptime(check_in_date, '%Y-%m-%d').date()
                except:
                    pass
            
            check_out_date = request.form.get('check_out_date', '').strip()
            if check_out_date:
                try:
                    student.check_out_date = datetime.strptime(check_out_date, '%Y-%m-%d').date()
                except:
                    pass
            
            residence_permit_expiry = request.form.get('residence_permit_expiry', '').strip()
            if residence_permit_expiry:
                try:
                    student.residence_permit_expiry = datetime.strptime(residence_permit_expiry, '%Y-%m-%d').date()
                except:
                    pass
            
            count += 1
    
    db.session.commit()
    
    flash(f'已成功修改 {count} 名学生的信息', 'success')
    return redirect(url_for('students.index'))


@bp.route('/detail/<int:id>', methods=['GET', 'POST'])
@login_required
def detail(id):
    """学生详情"""
    student = Student.query.get_or_404(id)
    
    # 处理缴费提交
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'payment':
            payment_amount = request.form.get('payment_amount', '0', type=float)
            new_due_date = request.form.get('new_due_date')
            payment_method = request.form.get('payment_method', '现金')
            
            # 更新已缴金额
            if payment_amount and payment_amount > 0:
                student.total_paid = (student.total_paid or 0) + payment_amount
                
                # 添加缴费记录
                fee_record = FeeRecord()
                fee_record.student_id = student.id
                fee_record.amount = payment_amount
                fee_record.payment_date = date.today()
                fee_record.payment_method = payment_method
                fee_record.notes = f'通过学生详情页缴纳'
                db.session.add(fee_record)
            
            # 更新到期日期
            if new_due_date:
                try:
                    student.payment_due_date = datetime.strptime(new_due_date, '%Y-%m-%d').date()
                except:
                    pass
            
            db.session.commit()
            flash(f'缴费成功！已缴纳 ¥{payment_amount:.2f}', 'success')
            return redirect(url_for('students.detail', id=id))
    
    fee_records = FeeRecord.query.filter_by(student_id=id).order_by(FeeRecord.payment_date.desc()).all()
    
    return render_template('students/detail.html',
                         title=f'{student.name} - 详情',
                         student=student,
                         fee_records=fee_records)


@bp.route('/batch-import', methods=['GET', 'POST'])
@login_required
@permission_required('write')
def batch_import():
    """批量导入学生"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('请选择要上传的文件', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('请选择要上传的文件', 'danger')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('只支持 Excel 文件 (.xlsx, .xls)', 'danger')
            return redirect(request.url)
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1] if cell.value]
            
            success_count = 0
            error_count = 0
            
            def parse_date(date_value):
                """解析日期，支持字符串和datetime对象"""
                if not date_value:
                    return None
                if isinstance(date_value, datetime):
                    return date_value.date()
                if isinstance(date_value, date):
                    return date_value
                # 字符串格式 YYYY-MM-DD 或 YYYY/MM/DD
                date_str = str(date_value).strip()
                try:
                    return datetime.strptime(date_str, '%Y-%m-%d').date()
                except:
                    try:
                        return datetime.strptime(date_str, '%Y/%m/%d').date()
                    except:
                        return None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # 姓名为空则跳过
                    continue
                
                row_data = dict(zip(headers, row))
                
                try:
                    student = Student()
                    
                    # 基本信息
                    student.name = str(row_data.get('姓名', '')).strip()
                    student.gender = str(row_data.get('性别', '')).strip() if row_data.get('性别') else None
                    student.nationality = str(row_data.get('国籍', '')).strip() if row_data.get('国籍') else None
                    student.department = str(row_data.get('所属业务部', '')).strip() if row_data.get('所属业务部') else None
                    student.major = str(row_data.get('专业', '')).strip() if row_data.get('专业') else None
                    
                    # 房间类型处理
                    room_type = row_data.get('房间类型', '')
                    if room_type and '单人间' in str(room_type):
                        student.bed_occupancy = 2  # 单人间占用2个床位
                    else:
                        student.bed_occupancy = 1  # 双人间占用1个床位
                    
                    # 处理楼栋号和房间号
                    building = row_data.get('楼栋号', '')
                    room_number = row_data.get('房间号', '')
                    
                    if building and room_number:
                        building = str(building).strip()
                        room_number = str(room_number).strip()
                        # 查找对应房间
                        room = Room.query.filter_by(building=building, room_number=room_number).first()
                        if room:
                            # 检查房间容量是否足够
                            if room.current_occupancy + student.bed_occupancy <= room.capacity:
                                student.room_id = room.id
                                room.current_occupancy += student.bed_occupancy
                                if room.current_occupancy >= room.capacity:
                                    room.status = 'full'
                            else:
                                # 房间容量不足，记录警告但不分配
                                pass
                    
                    student.check_in_date = parse_date(row_data.get('入住日期'))
                    student.check_out_date = parse_date(row_data.get('预计离开日期'))
                    student.payment_due_date = parse_date(row_data.get('缴费到期日期'))
                    student.residence_permit_expiry = parse_date(row_data.get('居留许可到期日期'))
                    
                    # 处理收费标准
                    fee_standard_name = row_data.get('收费标准', '')
                    if fee_standard_name:
                        fee_standard = FeeStandard.query.filter_by(name=str(fee_standard_name).strip(), is_active=True).first()
                        if fee_standard:
                            student.fee_standard_id = fee_standard.id
                    
                    # 处理本次缴纳房费
                    current_payment = row_data.get('本次缴纳房费', 0)
                    payment_amount_to_record = 0
                    if current_payment:
                        try:
                            payment_amount_to_record = float(current_payment)
                            if payment_amount_to_record > 0:
                                student.total_paid = payment_amount_to_record
                        except:
                            payment_amount_to_record = 0
                    
                    student.notes = str(row_data.get('备注', '')).strip() if row_data.get('备注') else None
                    student.status = 'active'
                    
                    db.session.add(student)
                    db.session.flush()  # 获取student.id
                    
                    # 创建缴费记录（在flush之后，student.id已生成）
                    if payment_amount_to_record > 0:
                        fee_record = FeeRecord(
                            student_id=student.id,
                            amount=payment_amount_to_record,
                            payment_date=date.today(),
                            payment_method='批量导入',
                            notes='批量导入时录入'
                        )
                        db.session.add(fee_record)
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    continue
            
            db.session.commit()
            flash(f'导入完成！成功: {success_count} 条，失败: {error_count} 条', 'success')
            return redirect(url_for('students.index'))
            
        except Exception as e:
            flash(f'导入失败: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('students/batch_import.html', title='批量导入学生')


@bp.route('/undo-recent-import', methods=['GET', 'POST'])
@login_required
def undo_recent_import():
    """撤销最近批量导入的学生
    
    删除最近1小时内创建的学生记录
    """
    if request.method == 'POST':
        # 获取最近1小时内创建的学生
        from datetime import datetime, timedelta
        cutoff_time = datetime.utcnow() - timedelta(hours=1)
        
        # 查询最近创建的学生（未分配房间的优先，避免影响房间状态）
        recent_students = Student.query.filter(
            Student.created_at >= cutoff_time,
            Student.status == 'active'
        ).all()
        
        if not recent_students:
            flash('没有找到最近导入的学生记录', 'warning')
            return redirect(url_for('students.index'))
        
        count = 0
        for student in recent_students:
            # 解除房间关联（如果有的话）
            if student.room_id:
                # 更新房间占用
                room = Room.query.get(student.room_id)
                if room:
                    room.current_occupancy -= student.bed_occupancy
                student.room_id = None
            
            # 删除学生的缴费记录
            FeeRecord.query.filter_by(student_id=student.id).delete()
            
            # 删除学生
            db.session.delete(student)
            count += 1
        
        db.session.commit()
        flash(f'已撤销最近导入的 {count} 名学生', 'success')
        return redirect(url_for('students.index'))
    
    # GET请求 - 显示确认页面
    from datetime import datetime, timedelta
    cutoff_time = datetime.utcnow() - timedelta(hours=1)
    
    recent_students = Student.query.filter(
        Student.created_at >= cutoff_time,
        Student.status == 'active'
    ).order_by(Student.created_at.desc()).limit(100).all()
    
    return render_template('students/undo_import.html', 
                         title='撤销最近导入',
                         students=recent_students)


@bp.route('/<int:id>/fees')
@login_required
def student_fees(id):
    """学生缴费记录"""
    student = Student.query.get_or_404(id)
    fee_records = FeeRecord.query.filter_by(student_id=id).order_by(FeeRecord.payment_date.desc()).all()
    
    return render_template('students/student_fees.html', 
                         title=f'{student.name} - 缴费记录',
                         student=student,
                         fee_records=fee_records)


@bp.route('/export-template')
@login_required
def export_template():
    """下载学生导入模板（带下拉菜单）
    
    包含添加学生页面的所有字段（除已缴房费总计）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    
    wb = Workbook()
    ws = wb.active
    ws.title = '学生导入模板'
    
    # 表头 - 包含添加学生页面的所有字段
    headers = [
        '姓名', '性别', '国籍', '护照号码', '专业', 
        '楼栋号', '房间号', '房间类型', 
        '入住日期', '预计离开日期', 
        '收费标准', '本次缴纳房费', '缴费到期日期', 
        '居留许可到期日期', '备注'
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 设置列宽
    column_widths = {
        'A': 12,  # 姓名
        'B': 8,   # 性别
        'C': 12,  # 国籍
        'D': 15,  # 所属业务部
        'E': 18,  # 专业
        'F': 10,  # 楼栋号
        'G': 10,  # 房间号
        'H': 12,  # 房间类型
        'I': 14,  # 入住日期
        'J': 14,  # 预计离开日期
        'K': 15,  # 收费标准
        'L': 14,  # 本次缴纳房费
        'M': 14,  # 缴费到期日期
        'N': 16,  # 居留许可到期日期
        'O': 25,  # 备注
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    
    # 获取系统数据用于下拉菜单
    from app.models import Room, FeeStandard
    
    # 楼栋号列表
    rooms = Room.query.filter(Room.status != 'archived').all()
    building_list = sorted(set(r.building for r in rooms if r.building))
    room_number_list = sorted(set(r.room_number for r in rooms if r.room_number))
    
    # 收费标准列表
    fee_standards = FeeStandard.query.filter_by(is_active=True).all()
    fee_list = [fs.name for fs in fee_standards if fs.name]
    
    # 1. 性别下拉菜单 (B列)
    gender_dv = DataValidation(
        type="list",
        formula1='"男,女,其他"',
        allow_blank=True
    )
    gender_dv.error = '请从下拉列表中选择性别'
    gender_dv.errorTitle = '无效的性别'
    gender_dv.prompt = '请选择性别'
    gender_dv.promptTitle = '性别'
    ws.add_data_validation(gender_dv)
    gender_dv.add('B2:B1000')
    
    # 2. 楼栋号下拉菜单 (F列)
    if building_list:
        building_options = ','.join(building_list)
        building_dv = DataValidation(
            type="list",
            formula1=f'"{building_options}"',
            allow_blank=True
        )
        building_dv.error = '请从下拉列表中选择楼栋号'
        building_dv.errorTitle = '无效的楼栋号'
        building_dv.prompt = '请选择楼栋号'
        building_dv.promptTitle = '楼栋号'
        ws.add_data_validation(building_dv)
        building_dv.add('F2:F1000')
    
    # 3. 房间号下拉菜单 (G列)
    if room_number_list:
        room_options = ','.join(room_number_list)
        room_dv = DataValidation(
            type="list",
            formula1=f'"{room_options}"',
            allow_blank=True
        )
        room_dv.error = '请从下拉列表中选择房间号'
        room_dv.errorTitle = '无效的房间号'
        room_dv.prompt = '请选择房间号'
        room_dv.promptTitle = '房间号'
        ws.add_data_validation(room_dv)
        room_dv.add('G2:G1000')
    
    # 4. 房间类型下拉菜单 (H列)
    room_type_dv = DataValidation(
        type="list",
        formula1='"双人间（占1床位）,单人间（占2床位）"',
        allow_blank=True
    )
    room_type_dv.error = '请从下拉列表中选择房间类型'
    room_type_dv.errorTitle = '无效的房间类型'
    room_type_dv.prompt = '单人间独享整个房间'
    room_type_dv.promptTitle = '房间类型'
    ws.add_data_validation(room_type_dv)
    room_type_dv.add('H2:H1000')
    
    # 5. 收费标准下拉菜单 (K列)
    if fee_list:
        fee_options = ','.join(fee_list)
        fee_dv = DataValidation(
            type="list",
            formula1=f'"{fee_options}"',
            allow_blank=True
        )
        fee_dv.error = '请从下拉列表中选择收费标准'
        fee_dv.errorTitle = '无效的收费标准'
        fee_dv.prompt = '请选择收费标准'
        fee_dv.promptTitle = '收费标准'
        ws.add_data_validation(fee_dv)
        fee_dv.add('K2:K1000')
    
    # 添加示例数据
    today = date.today().strftime('%Y-%m-%d')
    sample_data = [
        [
            '张三', '男', '中国', '', '计算机科学与技术',
            building_list[0] if building_list else '',
            room_number_list[0] if room_number_list else '',
            '双人间（占1床位）',
            today, '',
            fee_list[0] if fee_list else '', '0', '',
            '', ''
        ],
        [
            '李四', '女', '美国', 'P1234567', '软件工程',
            '', '', '单人间（占2床位）',
            today, '',
            fee_list[0] if fee_list else '', '0', '',
            '', ''
        ],
    ]
    
    # 设置数据行样式
    data_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border
            if row_idx == 2:  # 第一行示例数据高亮
                cell.fill = data_fill
    
    # 添加说明行
    note_row = len(sample_data) + 3
    ws.cell(row=note_row, column=1, value='填写说明：')
    ws.cell(row=note_row, column=1).font = Font(bold=True, color='FF0000')
    
    notes = [
        '1. 姓名为必填项',
        '2. 日期格式：YYYY-MM-DD（如 2026-04-16）',
        '3. 房间类型：双人间占用1个床位，单人间占用2个床位（独享房间）',
        '4. 本次缴纳房费：首次缴纳金额，如不缴纳填0或不填',
        '5. 缴费到期日期：根据收费标准计算，或手动填写',
        '6. 居留许可到期日期：外国学生必填，用于到期提醒',
    ]
    
    for i, note in enumerate(notes):
        ws.cell(row=note_row + 1 + i, column=1, value=note)
        ws.merge_cells(start_row=note_row + 1 + i, start_column=1, end_row=note_row + 1 + i, end_column=8)
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='学生导入模板.xlsx'
    )


@bp.route('/archived')
@login_required
def archived():
    """归档学生列表（已删除的学生，保留3年）"""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    search = request.args.get('search', '')
    
    # 查询已归档的学生
    query = Student.query.filter(Student.status == 'archived')
    
    if search:
        query = query.filter(
            (Student.name.contains(search)) |
            (Student.student_id.contains(search)) |
            (Student.department.contains(search))
        )
    
    pagination = query.order_by(Student.deleted_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    students = pagination.items
    
    return render_template('students/archived.html', 
                         title='归档学生',
                         students=students,
                         pagination=pagination,
                         search=search)


@bp.route('/export')
@login_required
def export():
    """导出当前学生列表（支持筛选条件）"""
    search = request.args.get('search', '')
    major = request.args.get('major', '')
    filter_status = request.args.get('filter', 'all')
    
    # 构建查询（与 index 视图相同的筛选逻辑）
    query = Student.query.filter(Student.status != 'archived')
    
    if search:
        query = query.filter(
            (Student.name.contains(search)) |
            (Student.student_id.contains(search)) |
            (Student.phone.contains(search))
        )
    
    if major:
        query = query.filter(Student.major.contains(major))
    
    if filter_status == 'housed':
        query = query.filter(Student.room_id != None)
    elif filter_status == 'unhoused':
        query = query.filter(Student.room_id == None)
    
    students = query.order_by(Student.created_at.desc()).all()
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '学生列表'
    
    # 设置样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['专业', '姓名', '性别', '国籍', '楼栋', '房间号', '房型', 
               '入住日期', '预计离开日期', '居留许可有效期', '所属业务部', 
               '学号', '联系电话', '欠费状态']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 数据行
    for row_idx, student in enumerate(students, 2):
        # 房型
        room_type = ''
        if student.room_id:
            room_type = '单人间' if student.bed_occupancy == 2 else '双人间'
        
        # 欠费状态
        has_arrears = student.has_arrears()
        arrears_status = f'欠费 ¥{student.calculate_arrears():.2f}' if has_arrears else '正常'
        
        # 房间信息
        building = student.room.building if student.room else ''
        room_number = student.room.room_number if student.room else ''
        
        row_data = [
            student.major or '',
            student.name or '',
            student.gender or '',
            student.nationality or '',
            building,
            room_number,
            room_type,
            student.check_in_date.strftime('%Y-%m-%d') if student.check_in_date else '',
            student.check_out_date.strftime('%Y-%m-%d') if student.check_out_date else '',
            student.residence_permit_expiry.strftime('%Y-%m-%d') if student.residence_permit_expiry else '',
            student.department or '',
            student.student_id or '',
            student.phone or '',
            arrears_status
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border
            # 欠费行高亮
            if col_idx == 14 and has_arrears:
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
    
    # 设置列宽
    column_widths = [20, 12, 8, 12, 10, 10, 10, 14, 14, 16, 15, 15, 15, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    
    # 生成文件名（包含日期时间）
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    download_name = f'学生列表_{now}.xlsx'
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )


@bp.route('/archived/export')
@login_required
def export_archived():
    """导出所有归档学生"""
    search = request.args.get('search', '')
    
    # 查询已归档的学生
    query = Student.query.filter(Student.status == 'archived')
    
    if search:
        query = query.filter(
            (Student.name.contains(search)) |
            (Student.student_id.contains(search)) |
            (Student.department.contains(search))
        )
    
    students = query.order_by(Student.deleted_at.desc()).all()
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '归档学生'
    
    # 设置样式
    header_fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['姓名', '性别', '国籍', '所属业务部', '专业', '学号', 
               '联系电话', '删除时间', '保留截止', '状态']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 数据行
    for row_idx, student in enumerate(students, 2):
        # 状态
        if student.retention_until and student.retention_until < date.today():
            status = '已过期'
            status_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        else:
            status = '已归档'
            status_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        row_data = [
            student.name or '',
            student.gender or '',
            student.nationality or '',
            student.department or '',
            student.major or '',
            student.student_id or '',
            student.phone or '',
            student.deleted_at.strftime('%Y-%m-%d %H:%M') if student.deleted_at else '',
            student.retention_until.strftime('%Y-%m-%d') if student.retention_until else '',
            status
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border
            # 过期行高亮
            if col_idx == 10 and student.retention_until and student.retention_until < date.today():
                cell.fill = status_fill
    
    # 设置列宽
    column_widths = [12, 8, 12, 15, 20, 15, 15, 18, 14, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    
    # 生成文件名（包含日期时间）
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    download_name = f'归档学生_{now}.xlsx'
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )


@bp.route('/arrears')
@login_required
def arrears():
    """欠费学生列表"""
    all_students = Student.query.filter(Student.status == 'active').all()
    arrears_students = [s for s in all_students if s.has_arrears()]
    
    # 按欠费金额排序
    arrears_students.sort(key=lambda s: s.calculate_arrears(), reverse=True)
    
    return render_template('students/arrears.html',
                         title='欠费学生',
                         arrears_students=arrears_students)
